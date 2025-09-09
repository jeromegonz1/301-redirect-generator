#!/usr/bin/env python3
"""
Tests TDD pour l'export XLS format Balt (Sprint 3.5)

Spécifications Balt:
- Format XLS (.xlsx) 
- 2 colonnes: Source, Target
- URLs absolues uniquement
- Pas de commentaires, confidence ou type redirection
- UTF-8, nom: redirections_balt.xlsx
"""

import pytest
import tempfile
import os
import sys
from pathlib import Path
from unittest.mock import patch, MagicMock

# Ajout du path source
sys.path.append(str(Path(__file__).parent.parent / "src"))

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    pytest.skip("openpyxl non installé", allow_module_level=True)


class TestBaltExport:
    """Tests TDD pour l'export XLS format Balt"""

    def test_generate_balt_xlsx_structure(self):
        """Test 1: Structure fichier XLS avec exactement 2 colonnes Source/Target"""
        from advanced_interface import generate_balt_xlsx
        
        # Données test avec URLs absolues
        redirects_data = [
            {
                'source': 'https://www.campinglescascades.com/fr/page1/',
                'target': 'https://www.les-cascades.com/fr/page1/',
                'type': '301',
                'comment': 'Correspondance exacte (confiance: 0.95)'
            },
            {
                'source': 'https://www.campinglescascades.com/en/page2/', 
                'target': 'https://www.les-cascades.com/en/',
                'type': '302',
                'comment': 'Fallback 302 vers la home EN'
            }
        ]
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            try:
                # Génération du fichier XLS
                file_path = generate_balt_xlsx(redirects_data, tmp_file.name)
                
                # Vérification existence fichier
                assert os.path.exists(file_path)
                
                # Chargement et vérification structure
                workbook = openpyxl.load_workbook(file_path)
                worksheet = workbook.active
                
                # Vérification en-têtes (ligne 1)
                assert worksheet['A1'].value == 'Source'
                assert worksheet['B1'].value == 'Target'
                
                # Vérification données ligne 2
                assert worksheet['A2'].value == 'https://www.campinglescascades.com/fr/page1/'
                assert worksheet['B2'].value == 'https://www.les-cascades.com/fr/page1/'
                
                # Vérification données ligne 3 (302 inclus)
                assert worksheet['A3'].value == 'https://www.campinglescascades.com/en/page2/'
                assert worksheet['B3'].value == 'https://www.les-cascades.com/en/'
                
                # Vérification: SEULEMENT 2 colonnes (pas de colonnes C, D, etc.)
                assert worksheet['C1'].value is None
                assert worksheet['C2'].value is None
                
                workbook.close()
                
            finally:
                # Nettoyage fichier temporaire
                if os.path.exists(tmp_file.name):
                    os.unlink(tmp_file.name)

    def test_generate_balt_xlsx_empty_data(self):
        """Test 2: Gestion des données vides (headers uniquement)"""
        from advanced_interface import generate_balt_xlsx
        
        redirects_data = []  # Aucune redirection
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            try:
                file_path = generate_balt_xlsx(redirects_data, tmp_file.name)
                
                workbook = openpyxl.load_workbook(file_path)
                worksheet = workbook.active
                
                # Headers présents même sans données
                assert worksheet['A1'].value == 'Source'
                assert worksheet['B1'].value == 'Target'
                
                # Ligne 2 vide (pas de données)
                assert worksheet['A2'].value is None
                assert worksheet['B2'].value is None
                
                workbook.close()
                
            finally:
                if os.path.exists(tmp_file.name):
                    os.unlink(tmp_file.name)

    def test_generate_balt_xlsx_mixed_redirects(self):
        """Test 3: Mélange 301 et 302, tous inclus dans export Balt"""
        from advanced_interface import generate_balt_xlsx
        
        redirects_data = [
            {'source': 'https://old.com/301-page', 'target': 'https://new.com/301-page', 'type': '301'},
            {'source': 'https://old.com/302-page', 'target': 'https://new.com/fr/', 'type': '302'},
            {'source': 'https://old.com/autre-301', 'target': 'https://new.com/autre-301', 'type': '301'}
        ]
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            try:
                file_path = generate_balt_xlsx(redirects_data, tmp_file.name)
                
                workbook = openpyxl.load_workbook(file_path)
                worksheet = workbook.active
                
                # Vérification: toutes les redirections présentes (301 ET 302)
                assert worksheet.max_row == 4  # 1 header + 3 données
                
                # Vérification données ligne par ligne
                assert worksheet['A2'].value == 'https://old.com/301-page'
                assert worksheet['B2'].value == 'https://new.com/301-page'
                
                assert worksheet['A3'].value == 'https://old.com/302-page'  
                assert worksheet['B3'].value == 'https://new.com/fr/'
                
                assert worksheet['A4'].value == 'https://old.com/autre-301'
                assert worksheet['B4'].value == 'https://new.com/autre-301'
                
                workbook.close()
                
            finally:
                if os.path.exists(tmp_file.name):
                    os.unlink(tmp_file.name)

    def test_generate_balt_xlsx_filename_default(self):
        """Test 4: Nom de fichier par défaut 'redirections_balt.xlsx'"""
        from advanced_interface import generate_balt_xlsx
        
        redirects_data = [
            {'source': 'https://test.com/page', 'target': 'https://new.com/page', 'type': '301'}
        ]
        
        with tempfile.TemporaryDirectory() as tmp_dir:
            # Sans spécifier de nom (utilise le défaut)
            file_path = generate_balt_xlsx(redirects_data, output_dir=tmp_dir)
            
            expected_path = os.path.join(tmp_dir, 'redirections_balt.xlsx')
            assert file_path == expected_path
            assert os.path.exists(expected_path)

    def test_balt_export_integration_streamlit(self):
        """Test 5: Intégration avec interface Streamlit (mock)"""
        # Ce test vérifiera l'intégration une fois implémentée dans advanced_interface.py
        
        # Mock des redirections filtrées comme dans l'interface réelle
        filtered_redirects = [
            {
                'source': 'https://www.campinglescascades.com/de/actualites-de/camping-4-sterne-wache/',
                'target': 'https://www.les-cascades.com/les-cascades/de/das-wichtigste-fuer-einen-gelungenen-aufenthalt-auf-dem-campingplatz-les-cascades/',
                'type': '301',
                'comment': '# Correspondance thématique évidente (confiance: 0.95)'
            }
        ]
        
        # Test que les données sont bien formatées pour l'export Balt
        assert len(filtered_redirects) > 0
        assert 'source' in filtered_redirects[0]
        assert 'target' in filtered_redirects[0]
        assert filtered_redirects[0]['source'].startswith('https://')
        assert filtered_redirects[0]['target'].startswith('https://')